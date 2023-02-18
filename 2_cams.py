import datetime
import cv2
import numpy as np
import os
import time
# from imutils.video import FPS
import pandas as pd
from openpyxl import load_workbook

def read_config_cams():
    # читаем конфиг
    f = open('config.ini', 'r')
    for line in f:
        l = [line.strip() for line in f]
    f.close()
    input1 = f'{l[0]}'
    input2 = f'{l[1]}'
    return input1, input2

# счётчик FPS
def set_FPS():
    pTime = 0
    cTime = time.time()
    fps = 1 / (cTime - pTime)
    pTime = cTime
    cv2.putText(imin, "FPS:" + str(round(fps)), (10, 40), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 1)


# Создаем БД в эксель по умолчанию
default_dir_db = 'db'


# каталог текущего дня ввиде даты 2022-MM-DD
sheet_name1 = datetime.datetime.now()
sheet_name1 = sheet_name1.strftime("%Y-%m-%d")
os.makedirs(f'{default_dir_db}/{sheet_name1}', exist_ok=True)
db_default = 'db.xlsx'
# Создаем пустой датафрейм за текущую дату если отсутствует
check_file = os.path.exists(f'{default_dir_db}/{sheet_name1}/db.xlsx')
person_in_fio = 'Не определено'
person_in_time = 'Время не определено'
person_out_fio = 'Не определено'
person_out_time = 'Время не определено'

if (check_file == False):
    db_1 = pd.DataFrame(columns=['id_dt', 'event', 'id_fio', 'fio', 'fio_path', 'status'])
    # Создаем пустой excel на текущую дату если она не существует
    db_1.to_excel(f'{default_dir_db}/{sheet_name1}/db.xlsx', sheet_name=sheet_name1, index=False)
    temp_db_in = db_1.copy()
    #temp_db_in.drop(temp_db_in.index, inplace=True)
    temp_db_out = db_1.copy()
    #temp_db_out.drop(temp_db_out.index, inplace=True)
else:
    db_1 = pd.read_excel(f'{default_dir_db}/{sheet_name1}/db.xlsx', sheet_name=sheet_name1)
    filter_in = db_1['status'] == 'in'
    temp_db_in = db_1.loc[filter_in].copy()
    filter_out = db_1['status'] == 'out'
    temp_db_out = db_1.loc[filter_out].copy()
    if len(temp_db_in.index)>0:
        person_in_fio = temp_db_in['fio'].where(temp_db_in['id_dt'] == temp_db_in['id_dt'].max()).dropna().values[0]
        person_in_time = temp_db_in['id_dt'].where(temp_db_in['id_dt'] == temp_db_in['id_dt'].max()).dropna().tolist()
        person_in_time = person_in_time[0]
        person_in_time = datetime.datetime.strftime(person_in_time, '%Y-%m-%d %H:%M:%S')
        temp_db_in.drop(temp_db_in.index, inplace=True)

    if len(temp_db_out.index)>0:
        person_out_fio = temp_db_out['fio'].where(temp_db_out['id_dt'] == temp_db_out['id_dt'].max()).dropna().values[0]
        person_out_time = temp_db_out['id_dt'].where(temp_db_out['id_dt'] == temp_db_out['id_dt'].max()).dropna().tolist()
        person_out_time = person_out_time[0]
        person_out_time = datetime.datetime.strftime(person_out_time, '%Y-%m-%d %H:%M:%S')
        temp_db_out.drop(temp_db_out.index, inplace=True)

#print(person_out_time)

haar_file = 'haarcascade_frontalface_default.xml'
datasets = 'datasets'

(images, lables, names, id) = ([], [], {}, 0)
for (subdirs, dirs, files) in os.walk(datasets):
    for subdir in dirs:
        names[id] = subdir
        subjectpath = os.path.join(datasets, subdir)
        for filename in os.listdir(subjectpath):
            path = subjectpath + '/' + filename
            lable = id
            images.append(cv2.imread(path, 0))
            lables.append(int(lable))
        id += 1
        (width, height) = (130, 100)
(images, lables) = [np.array(lis) for lis in [images, lables]]
model = cv2.face.LBPHFaceRecognizer_create()
model.train(images, lables)
face_cascade = cv2.CascadeClassifier(haar_file)

try:
    # webcam = cv2.VideoCapture('rtsp://admin:Asaaa!@@192.168.0.109/cam/realmonitor?channel=1&subtype=0')

    webcamin = cv2.VideoCapture(read_config_cams()[0])
    webcamout = cv2.VideoCapture(read_config_cams()[1])
except:
    print('Не удалось запустить поток')
    quit()

while True:

    (_, imin) = webcamin.read()
    #set_FPS()

    (_, imout) = webcamout.read()

    grayin = cv2.cvtColor(imin, cv2.COLOR_BGR2GRAY)
    grayout = cv2.cvtColor(imout, cv2.COLOR_BGR2GRAY)

    facesin = face_cascade.detectMultiScale(grayin, 1.3, 4)
    facesout = face_cascade.detectMultiScale(grayout, 1.3, 4)
    #print(facesout)

    # hsvin = cv2.cvtColor(imin, cv2.COLOR_BGR2HSV)
    # hsvout = cv2.cvtColor(imout, cv2.COLOR_BGR2HSV)
    #lower_blue = np.array([110, 50, 50])
    #upper_blue = np.array([130, 255, 255])
  
    for (x, y, w, h) in facesin:
        cv2.rectangle(imin, (x, y), (x + w, y + h), (255, 0, 0), 2)
        facein = grayin[y:y + h, x:x + w]
        face_resize_in = cv2.resize(facein, (width, height))
        predictionin = model.predict(face_resize_in)
        cv2.rectangle(imin, (x, y), (x + w, y + h), (0, 255, 0), 3)

        if predictionin[1] < 90:
            cv2.putText(imin, 'The person of % s - %.0f' % (names[predictionin[0]], predictionin[1]), (x - 10, y - 10),
                        cv2.FONT_HERSHEY_PLAIN, 1, (0, 0, 255), 2)
            now1 = datetime.datetime.now()
            #now1=datetime.datetime.strftime(now1, '%Y-%m-%d %H:%M:%S')
            id_dt_in = now1
            event_in = 'Enter'
            id_fio_in = predictionin[0]
            #print(predictionin[1])
            fio_in = f'{names[predictionin[0]]}'
            fio_path = f'datasets/{names[predictionin[0]]}'
            status = 'in'
            # print(id_dt_in, event_in, id_fio_in, fio_in, fio_path, status)
            new_row = {'id_dt': now1,
                       'event': event_in,
                       'id_fio': id_fio_in,
                       'fio': fio_in,
                       'fio_path': fio_path,
                       'status': status
                       }
            temp_db_in.loc[len(temp_db_in.index)] = new_row
            person_in_fio = temp_db_in['fio'].where(temp_db_in['id_dt'] == temp_db_in['id_dt'].max()).dropna().values[0]
            person_in_time = temp_db_in['id_dt'].\
                where(temp_db_in['id_dt'] == temp_db_in['id_dt'].max()).dropna().tolist()
            person_in_time = person_in_time[0]
            person_in_time = datetime.datetime.strftime(person_in_time, '%Y-%m-%d %H:%M:%S')

    for (x, y, w, h) in facesout:
        cv2.rectangle(imout, (x, y), (x + w, y + h), (255, 0, 0), 2)
        faceout = grayout[y:y + h, x:x + w]
        face_resize_out = cv2.resize(faceout, (width, height))
        predictionout = model.predict(face_resize_out)
        cv2.rectangle(imout, (x, y), (x + w, y + h), (0, 0, 255), 3)
        if predictionout[1] < 90:
            cv2.putText(imout, ( 'The person of % s - %.0f' % (names[predictionout[0]], predictionout[1])),
                        (x - 10, y - 10),
                        cv2.FONT_HERSHEY_PLAIN, 1, (0, 0, 255), 2)
            now1 = datetime.datetime.now()
            # now1=datetime.datetime.strftime(now1, '%Y-%m-%d %H:%M:%S')
            id_dt_out = now1
            event_out = 'Exit'
            id_fio_out = predictionout[0]
            #print(predictionout[0])
            fio_out = f'{names[predictionout[0]]}'
            fio_path = f'datasets/{names[predictionout[0]]}'
            status = 'out'
            #print(id_dt_out, event_out, id_fio_out, fio_out, fio_path, status)
            new_row = {'id_dt': now1,
                       'event': event_out,
                       'id_fio': id_fio_out,
                       'fio': fio_out,
                       'fio_path': fio_path,
                       'status': status
                       }
            temp_db_out.loc[len(temp_db_out.index)] = new_row

            person_out_fio = \
            temp_db_out['fio'].where(temp_db_out['id_dt'] == temp_db_out['id_dt'].max()).dropna().values[0]
            person_out_time = temp_db_out['id_dt'].where(
                temp_db_out['id_dt'] == temp_db_out['id_dt'].max()).dropna().tolist()
            person_out_time = person_out_time[0]
            person_out_time = datetime.datetime.strftime(person_out_time, '%Y-%m-%d %H:%M:%S')





    info = [
        ("ФИО последнего вышедшего", person_out_fio),
        ("Время входа", person_out_time),
        ("ФИО последнего вошедшего", person_in_fio),
        ("Время выхода", person_in_time),
        ("Текущее время", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    ]
    for (i, (k, v)) in enumerate(info):
        text = "{}: {}".format(k, v)
        cv2.putText(imin, text, (8, ((i * 20) + 20)),
                    cv2.FONT_HERSHEY_COMPLEX, 0.6, (0, 255, 0), 2)
        cv2.putText(imout, text, (8, ((i * 20) + 20)),
                    cv2.FONT_HERSHEY_COMPLEX, 0.6, (0, 0, 255), 2)

    cv2.imshow('Livein', imin)
    cv2.imshow('LiveOut', imout)



    key = cv2.waitKey(1)
    if key == 27:
        # print(db_1)
        webcamin.release()
        webcamout.release()
        cv2.destroyAllWindows()
        filter_in = temp_db_in['status'] == 'in'
        temp_db_in = temp_db_in.loc[temp_db_in.groupby("id_fio")['id_dt'].idxmax()]

        filter_out = temp_db_out['status'] == 'out'
        temp_db_out = temp_db_out.loc[temp_db_out.groupby("id_fio")['id_dt'].idxmax()]

        col_names = ['id_dt_in', 'event_in', 'id_fio_in', 'fio_in', 'fio_path', 'status']
        # Сбор данных во временную бд
        temp_db_in.to_excel(f'{default_dir_db}/{sheet_name1}/temp_db_in.xlsx', sheet_name=sheet_name1, index=False)
        temp_db_out.to_excel(f'{default_dir_db}/{sheet_name1}/temp_db_out.xlsx', sheet_name=sheet_name1, index=False)

        book = load_workbook(f'{default_dir_db}/{sheet_name1}/db.xlsx')
        writer = pd.ExcelWriter(f'{default_dir_db}/{sheet_name1}/db.xlsx', engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}

        for sheetname in writer.sheets:
            temp_db_in.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index=False,
                             header=False)
        for sheetname in writer.sheets:
            temp_db_out.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index=False,
                             header=False)
        writer.save()
        temp_db_in.drop(temp_db_in.index, inplace=True)
        temp_db_out.drop(temp_db_out.index, inplace=True)
        os.remove(f'{default_dir_db}/{sheet_name1}/temp_db_in.xlsx')
        os.remove(f'{default_dir_db}/{sheet_name1}/temp_db_out.xlsx')
        break
